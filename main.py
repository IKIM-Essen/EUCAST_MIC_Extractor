import pandas as pd
from openpyxl import load_workbook
import re

def load_raw_workbook(path):
    # Load
    wb = load_workbook(path, rich_text=True)
    sheet_dfs = {}
    for sheet in wb.worksheets:
        data = []
        # Clean
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                if isinstance(cell.value, str):
                    cleaned_value = re.sub(r'<.*?>', '', cell.value)    # Remove HTML-like tags
                    cleaned_value = re.sub(r"[\']", "", cleaned_value)
                    row_data.append(cleaned_value)
                elif isinstance(cell.value, list):
                    cleaned_value = cell.value[0]
                    if cleaned_value == "":
                        cleaned_value = cell.value[1]
                    row_data.append(cleaned_value)
                else:
                    row_data.append(cell.value)  # Append non-string values as-is
            data.append(row_data)  # Append the processed row
        
        # Store
        df = pd.DataFrame(data)
        sheet_dfs[sheet.title] = df

    return sheet_dfs

def set_MIC_index(input_table):
    #Preprocessing
    input_table = input_table[input_table[0].notna()].copy()
    input_table.loc[:, 1] = input_table[1].astype(str)
    input_table["MIC_Index"] = input_table[1].str.find('MIC breakpoint')

    inside_mic = False
    next_is_header = False

    # Search for MIC
    for idx in input_table.index:
        entry = input_table.at[idx, "MIC_Index"]
        if entry == 0:  # Found the 'MIC breakpoint'
            inside_mic = True
            next_is_header = True
            continue
        if inside_mic:
            if next_is_header:
                next_is_header = False
                continue
            if entry == -1:  # No 'MIC breakpoint'
                input_table.at[idx, "MIC_Index"] = 1.0
            if pd.isna(input_table.at[idx, "MIC_Index"]): # Reset when a row is NaN (indicating end of MIC section)
                inside_mic = False

    return input_table

def shrink_clean_table(input_table):
        # Set up output
        header = {'Name': [], 'S <=': [], 'R >': []}
        output_table = pd.DataFrame(header)

        # Shrink
        for idx in input_table.index:
                if input_table["MIC_Index"][idx] == 1.0:
                        new_row = {"Name" : input_table[input_table.columns[0]][idx], "S <=" : input_table[input_table.columns[1]][idx], "R >" : input_table[input_table.columns[2]][idx]}
                        output_table = pd.concat([output_table, pd.DataFrame([new_row])], ignore_index=True)
                        
        # Clean table     
        output_table['Name'] = output_table['Name'].map(lambda x: re.sub(r",", "", x) if isinstance(x, str) else x)
        output_table['Name'] = output_table['Name'].map(lambda x: re.sub(r"\n", "", x) if isinstance(x, str) else x)
        output_table['S <='] = output_table['S <='].map(lambda x: re.sub(r"[()]", "", x) if isinstance(x, str) else x)
        output_table['R >'] = output_table['R >'].map(lambda x: re.sub(r"[()]", "", x) if isinstance(x, str) else x)

        return output_table

raw_workbook = load_raw_workbook('resources/v_14.0_Breakpoint_Tables.xlsx')
raw_table = raw_workbook["Pseudomonas"]
mic_table = set_MIC_index(raw_table)
shrunk_table = shrink_clean_table(mic_table)
pd.set_option('display.max_columns', 300)
pd.set_option('display.max_rows', 300)
print(shrunk_table)