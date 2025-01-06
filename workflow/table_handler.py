# Copyright 2025 by Miriam Balzer & Julian Welling, University of Duisburg-Essen
# Licensed under the MIT License
# This file may be copied, modified, and distributed under the terms of the MIT License.

from openpyxl import load_workbook
import pandas as pd
import re

column_name = 'MIC_Index'

def load_raw_table(path, sheet_name):
    # load workbook
    wb = load_workbook(path, rich_text=True)
    
    # Check if the sheet_name exists
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
    
    # Process only the matching sheet
    sheet = wb[sheet_name]
    data = []
    
    counter = 0
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            if isinstance(cell.value, str):
                cleaned_value = cell.value
                cleaned_value = re.sub(r'<.*?>', '', cleaned_value)    # Remove HTML-like tags
                cleaned_value = re.sub(r"[\']", "", cleaned_value)  # Remove '
                row_data.append(cleaned_value)
            elif isinstance(cell.value, list):
                cleaned_value = cell.value[0]
                if isinstance(cleaned_value, str) and cleaned_value == "":
                    cleaned_value = cell.value[1]
                row_data.append(str(cleaned_value))
            else:
                row_data.append(str(cell.value))  # Append non-string values as-is
        data.append(row_data)  # Append the processed row

    # Return the processed DataFrame for the matching sheet
    return pd.DataFrame(data)

def set_MIC_index(input_table):
    #Preprocessing
    input_table = input_table[input_table[0].notna()].copy()
    input_table.loc[:, 1] = input_table[1].astype(str)
    input_table[column_name] = input_table[1].str.find('MIC breakpoint')

    inside_mic = False
    next_is_header = False

    # Search for MIC
    for idx in input_table.index:
        entry = input_table.at[idx, column_name]
        if entry == 0:  # Found the 'MIC breakpoint'
            inside_mic = True
            continue
        if inside_mic:
            if entry == -1:  # No 'MIC breakpoint'
                input_table.at[idx, column_name] = 1.0
            if pd.isna(input_table.at[idx, column_name]): # Reset when a row is NaN (indicating end of MIC section)
                inside_mic = False

    return input_table

def shrink_clean_table(input_table):
    # Set up output
    header = {'Name': [], 'S <=': [], 'R >': []}
    output_table = pd.DataFrame(header)

    # Shrink
    for idx in input_table.index:
            if input_table[column_name][idx] == 1.0:
                    new_row = {"Name" : input_table[input_table.columns[0]][idx], "S <=" : input_table[input_table.columns[1]][idx], "R >" : input_table[input_table.columns[2]][idx]}
                    output_table = pd.concat([output_table, pd.DataFrame([new_row])], ignore_index=True)
                    
    # Clean table
    output_table = output_table.dropna(how='any')
    output_table = output_table[~output_table.isin(["None"]).any(axis=1)]
    output_table['Name'] = output_table['Name'].map(lambda x: re.sub(r",", "", x) if isinstance(x, str) else x)
    output_table['Name'] = output_table['Name'].map(lambda x: re.sub(r"\n", "", x) if isinstance(x, str) else x)
    output_table['S <='] = output_table['S <='].map(lambda x: re.sub(r"[()]", "", x) if isinstance(x, str) else x)
    output_table['R >'] = output_table['R >'].map(lambda x: re.sub(r"[()]", "", x) if isinstance(x, str) else x)

    return output_table
